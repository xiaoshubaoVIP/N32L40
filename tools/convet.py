import openpyxl
import Scanf
import os

path = "./1/"

sheet = ["sheet1", "sheet2", "sheet3", "sheet4", "sheet5"]
match_setp = ["co calibration", "co calibration setp %d start"]
match_co = ["co mvol", "co mvol:%d"]
match_avg = ["avg", "avg=%d"]

workbook = openpyxl.Workbook()
workbook.remove(workbook.active)

for s in sheet:
    workbook.create_sheet(s)

workbook["sheet5"].append(["0", "30", "150", "300"])

files = os.listdir(path)

for file in files:
    with open(path + file, "r") as f:
        setp=0
        co_lst=[]
        for l in f:
            if match_setp[0] in l:
                setp = Scanf.scanf(match_setp[1], l)
                co_lst=[]

            if match_co[0] in l:
                co_lst += Scanf.scanf(match_co[1], l)

            if match_avg[0] in l:
                avg = Scanf.scanf(match_avg[1], l)
                shee = workbook.active
                column = Scanf.scanf("%d.txt", file)
                for i in range(1, len(co_lst) + 1):
                    workbook["sheet" + str(setp[0])].cell(i, column[0] + 1, co_lst[i - 1])
                workbook["sheet5"].cell(column[0] + 2, setp[0], avg[0])

workbook.save("s.xlsx")
