import Scanf
import os
import openpyxl

workbook = openpyxl.Workbook()

def main():
    files = os.listdir()

    for file in files:
        if ".txt" in file:
            n = Scanf.scanf("%d.txt", file)
            workbook.create_sheet(str(n[0]))
            with open(file,"r") as file:
                conut = 0
                for i in range(0, 30000):
                    line = file.readline()
                    t = Scanf.scanf("D:%d, DVol:%f   L:%d, LVol:%f   L-D:%d, D_LVol:%f", line)
                    if t:
                        conut = conut + 1
                        workbook[str(n[0])].cell(conut, n[0], t[5])
    workbook.save("s.xlsx")
    return


if __name__ == "__main__":
    dir_name, file_name = os.path.split(__file__)
    os.chdir(dir_name)
    main()