import Scanf
import os
import openpyxl

workbook = openpyxl.Workbook()

def main():
    files = os.listdir()
    workbook.create_sheet("1")
    raw = 0
    column = 0

    for file in files:
        column = column + 1
        if ".txt" in file:
            with open(file,"r") as file:
                for line in file:
                    t = Scanf.scanf("co sample:%f(mv),b=%d,k=%d,ppm=%d", line)
                    if t:
                        raw = raw + 1
                        workbook["1"].cell(raw, column, t[3])
    workbook.save("s.xlsx")
    return


if __name__ == "__main__":
    dir_name, file_name = os.path.split(__file__)
    os.chdir(dir_name)
    main()